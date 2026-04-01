#!/usr/bin/env python3
"""
Dental Addict — Module Prospects v1
Scan hebdomadaire BCE Open Data (NACE 8623) + enrichissement email
"""

import csv, io, json, os, re, threading, time, zipfile
import urllib.request, urllib.error, urllib.parse
import gzip as gz
from datetime import datetime, timedelta
from html.parser import HTMLParser

# ── Configuration ─────────────────────────────────────────────
NACE_TARGETS = {"8623"}          # Pratique dentaire — priorité
STORAGE_FILE = "prospects_data.json"   # Fichier de persistance local

# Codes NACE connexes optionnels (commentés — activer si besoin)
# NACE_TARGETS |= {"3250","4646","8622"}

# ── Téléchargement BCE Open Data ──────────────────────────────
BCE_OPENDATA_BASE = "https://kbopub.economie.fgov.be/kbo-open-data/affiliation/public/files"

def get_bce_update_url():
    """
    Retourne l'URL du dernier fichier de modifications BCE Open Data.
    Format : KboOpenData_YYYY-MM-DD_update.zip
    Les fichiers sont publiés quotidiennement.
    On prend celui d'il y a 7 jours (semaine écoulée).
    """
    # Essayer les 7 derniers jours pour trouver un fichier disponible
    for days_ago in range(0, 10):
        d = datetime.now() - timedelta(days=days_ago)
        fname = f"KboOpenData_{d.strftime('%Y-%m-%d')}_update.zip"
        url   = f"{BCE_OPENDATA_BASE}/{fname}"
        try:
            req = urllib.request.Request(url, method="HEAD",
                headers={"User-Agent": "Mozilla/5.0"})
            urllib.request.urlopen(req, timeout=8)
            return url, fname, d.strftime("%Y-%m-%d")
        except Exception:
            continue
    return None, None, None


def download_and_parse_bce_update(credentials=None):
    """
    Télécharge le fichier de modifications BCE et extrait les nouvelles
    entreprises NACE 8623 créées dans la semaine.

    Le fichier ZIP contient plusieurs CSV :
    - enterprise.csv    : données entreprises
    - denomination.csv  : noms
    - address.csv       : adresses
    - activity.csv      : codes NACE (fichier clé pour filtrer)
    - contact.csv       : email, téléphone, site web

    Paramètre credentials : {"user": "...", "password": "..."}
    (requis si fichier protégé — inscription gratuite sur kbopub)
    """
    url, fname, date_str = get_bce_update_url()
    if not url:
        return {"ok": False, "error": "Aucun fichier de modifications trouvé (vérifiez la connexion)"}

    print(f"  Téléchargement : {fname}")

    try:
        # Authentification si credentials fournis
        if credentials:
            pm = urllib.request.HTTPPasswordMgrWithDefaultRealm()
            pm.add_password(None, BCE_OPENDATA_BASE, credentials["user"], credentials["password"])
            opener = urllib.request.build_opener(urllib.request.HTTPBasicAuthHandler(pm))
        else:
            opener = urllib.request.build_opener()

        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with opener.open(req, timeout=60) as r:
            zip_data = r.read()

        print(f"  ZIP téléchargé : {len(zip_data)//1024} KB")

    except urllib.error.HTTPError as e:
        if e.code == 401:
            return {"ok": False, "error": "Authentification requise — inscrivez-vous sur kbopub.economie.fgov.be/kbo-open-data"}
        if e.code == 403:
            return {"ok": False, "error": "Accès refusé — vérifiez vos credentials BCE Open Data"}
        return {"ok": False, "error": f"HTTP {e.code} lors du téléchargement"}
    except Exception as e:
        return {"ok": False, "error": f"Téléchargement échoué : {str(e)[:100]}"}

    # Parser le ZIP
    try:
        return _parse_bce_zip(zip_data, date_str)
    except Exception as e:
        return {"ok": False, "error": f"Erreur parsing ZIP : {str(e)[:100]}"}


def _parse_bce_zip(zip_data, date_str):
    """Parse le ZIP BCE et retourne les nouvelles entreprises NACE 8623."""
    with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
        files = zf.namelist()
        print(f"  Fichiers dans ZIP : {files}")

        # 1. Lire activity.csv → trouver les BCE avec NACE 8623
        bce_with_nace = set()
        nace_map = {}   # bce → liste de codes NACE
        activity_file = next((f for f in files if "activity" in f.lower()), None)
        if activity_file:
            with zf.open(activity_file) as af:
                reader = csv.DictReader(io.TextIOWrapper(af, encoding="utf-8", errors="replace"),
                                        delimiter=",")
                for row in reader:
                    code = row.get("NaceCode", row.get("ActivityCode", row.get("Code", ""))).strip()
                    ent  = row.get("EntityNumber", row.get("EnterpriseNumber", "")).strip().replace(".", "")
                    if not ent: continue
                    if ent not in nace_map: nace_map[ent] = []
                    nace_map[ent].append(code)
                    if code in NACE_TARGETS:
                        bce_with_nace.add(ent)

        print(f"  Entreprises NACE 8623 trouvées : {len(bce_with_nace)}")
        if not bce_with_nace:
            # Essayer avec le format du fichier insert
            activity_insert = next((f for f in files if "activity" in f.lower() and "insert" in f.lower()), None)
            if activity_insert:
                with zf.open(activity_insert) as af:
                    reader = csv.DictReader(io.TextIOWrapper(af, encoding="utf-8", errors="replace"))
                    for row in reader:
                        code = str(row.get("NaceCode","")).strip()
                        ent  = str(row.get("EntityNumber","")).strip().replace(".","")
                        if code in NACE_TARGETS and ent:
                            bce_with_nace.add(ent)

        # 2. Lire enterprise.csv → garder seulement les nouvelles créations
        enterprises = {}
        ent_file = next((f for f in files if "enterprise" in f.lower()), None)
        if ent_file:
            with zf.open(ent_file) as ef:
                reader = csv.DictReader(io.TextIOWrapper(ef, encoding="utf-8", errors="replace"))
                for row in reader:
                    ent = row.get("EnterpriseNumber", row.get("EntityNumber", "")).strip().replace(".", "")
                    if ent not in bce_with_nace: continue
                    # Garder uniquement les insertions (nouvelles créations)
                    situation = row.get("JuridicalSituation", "").strip()
                    status    = row.get("Status", "").strip()
                    start     = row.get("StartDate", "").strip()
                    # Filtrer : entreprises actives avec date de début récente
                    enterprises[ent] = {
                        "bce":              ent,
                        "forme_juridique":  row.get("TypeOfEnterprise", row.get("JuridicalForm", "")),
                        "situation":        situation,
                        "status":           status,
                        "date_creation":    start,
                        "nace_codes":       nace_map.get(ent, []),
                    }

        # 3. Lire denomination.csv → noms des entreprises
        denom_file = next((f for f in files if "denomination" in f.lower()), None)
        if denom_file:
            with zf.open(denom_file) as df:
                reader = csv.DictReader(io.TextIOWrapper(df, encoding="utf-8", errors="replace"))
                for row in reader:
                    ent  = row.get("EntityNumber", "").strip().replace(".", "")
                    if ent not in enterprises: continue
                    lang = row.get("Language", "").strip()
                    nom  = row.get("Denomination", "").strip()
                    # Préférer le nom en français, sinon premier disponible
                    if nom and ("nom" not in enterprises[ent] or lang == "FR"):
                        enterprises[ent]["nom"] = nom

        # 4. Lire address.csv → adresses
        addr_file = next((f for f in files if "address" in f.lower()), None)
        if addr_file:
            with zf.open(addr_file) as af:
                reader = csv.DictReader(io.TextIOWrapper(af, encoding="utf-8", errors="replace"))
                for row in reader:
                    ent  = row.get("EntityNumber", "").strip().replace(".", "")
                    if ent not in enterprises: continue
                    rue    = row.get("StreetFR", row.get("StreetNL", row.get("Street", ""))).strip()
                    num    = row.get("HouseNumber", "").strip()
                    cp     = row.get("Zipcode", row.get("ZipCode", "")).strip()
                    ville  = row.get("MunicipalityFR", row.get("MunicipalityNL", row.get("Municipality", ""))).strip()
                    adresse = " ".join(filter(None, [rue, num, cp, ville]))
                    if adresse:
                        enterprises[ent]["adresse"] = adresse
                        enterprises[ent]["cp"]       = cp
                        enterprises[ent]["ville"]    = ville

        # 5. Lire contact.csv → email, tél, website
        contact_file = next((f for f in files if "contact" in f.lower()), None)
        if contact_file:
            with zf.open(contact_file) as cf:
                reader = csv.DictReader(io.TextIOWrapper(cf, encoding="utf-8", errors="replace"))
                for row in reader:
                    ent   = row.get("EntityNumber", "").strip().replace(".", "")
                    if ent not in enterprises: continue
                    ctype = row.get("ContactType", row.get("TypeOfContact", "")).strip()
                    val   = row.get("Value", row.get("ContactValue", "")).strip()
                    if "EMAIL" in ctype.upper() and val and "@" in val:
                        enterprises[ent]["email"] = val
                    elif "WEB" in ctype.upper() or "URL" in ctype.upper():
                        enterprises[ent]["website"] = val[:80]
                    elif "TEL" in ctype.upper() or "PHONE" in ctype.upper():
                        enterprises[ent]["telephone"] = val

        result = list(enterprises.values())
        print(f"  {len(result)} nouveaux prospects extraits")
        return {
            "ok": True,
            "date": date_str,
            "count": len(result),
            "prospects": result
        }


# ── Enrichissement email depuis site web ─────────────────────
class EmailParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.emails = set()
    def handle_data(self, data):
        for m in re.finditer(r'[\w.\-+]+@[\w.\-]+\.[a-z]{2,6}', data, re.I):
            e = m.group(0).lower()
            # Filtrer les emails parasites
            if not any(x in e for x in ["example","noreply","no-reply","test@","@example",".png","@2x"]):
                self.emails.add(e)
    def handle_starttag(self, tag, attrs):
        if tag == "a":
            for name, val in attrs:
                if name == "href" and val and val.startswith("mailto:"):
                    email = val[7:].split("?")[0].strip()
                    if "@" in email: self.emails.add(email.lower())

def find_email_from_website(url):
    """Scrape le site web d'une entreprise pour trouver son email de contact."""
    if not url: return ""
    if not url.startswith("http"): url = "https://" + url
    try:
        req = urllib.request.Request(url.rstrip("/"), headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)",
            "Accept": "text/html", "Accept-Language": "fr-FR,fr;q=0.9"
        })
        with urllib.request.urlopen(req, timeout=8) as r:
            html = r.read().decode("utf-8", errors="replace")
        p = EmailParser(); p.feed(html)
        if p.emails: return sorted(p.emails)[0]

        # Essayer /contact
        contact_url = url.rstrip("/") + "/contact"
        req2 = urllib.request.Request(contact_url, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req2, timeout=6) as r2:
            html2 = r2.read().decode("utf-8", errors="replace")
        p2 = EmailParser(); p2.feed(html2)
        if p2.emails: return sorted(p2.emails)[0]
    except Exception:
        pass
    return ""


# ── Persistance JSON (stockage simple sur disque Render) ──────
def load_prospects():
    """Charge les prospects sauvegardés."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), STORAGE_FILE)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"scans": [], "prospects": []}

def save_prospects(data):
    """Sauvegarde les prospects."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), STORAGE_FILE)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def add_scan_result(scan_result, enrich_emails=True):
    """
    Intègre un scan BCE dans le stockage, avec enrichissement email optionnel.
    """
    store     = load_prospects()
    prospects = scan_result.get("prospects", [])
    date_str  = scan_result.get("date", datetime.now().strftime("%Y-%m-%d"))

    # Dédoublonner avec les prospects existants
    existing_bces = {p["bce"] for p in store["prospects"]}
    new_ones = [p for p in prospects if p["bce"] not in existing_bces]

    # Enrichissement email (si site web disponible)
    if enrich_emails:
        for p in new_ones:
            if not p.get("email") and p.get("website"):
                p["email"] = find_email_from_website(p["website"])
                time.sleep(0.3)  # politesse

    # Ajouter les champs manquants
    for p in new_ones:
        p.setdefault("nom",           "")
        p.setdefault("adresse",       "")
        p.setdefault("cp",            "")
        p.setdefault("ville",         "")
        p.setdefault("telephone",     "")
        p.setdefault("email",         "")
        p.setdefault("website",       "")
        p.setdefault("forme_juridique","")
        p.setdefault("date_creation", "")
        p.setdefault("statut_contact","Nouveau")  # Nouveau / Contacté / Sans suite
        p["scan_date"] = date_str

    store["prospects"].extend(new_ones)
    store["scans"].append({
        "date":      date_str,
        "new_count": len(new_ones),
        "total":     len(store["prospects"])
    })
    save_prospects(store)
    return {"new": len(new_ones), "total": len(store["prospects"])}


# ── Scheduler hebdomadaire ────────────────────────────────────
def start_weekly_scheduler(credentials=None):
    """Lance le scan tous les lundis matin à 8h00.
    Les credentials BCE sont lus depuis les variables d'environnement :
    BCE_USER et BCE_PASSWORD (à configurer dans Render → Environment).
    """
    def _loop():
        while True:
            now = datetime.now()
            days_until_monday = (7 - now.weekday()) % 7 or 7
            next_run = now.replace(hour=8, minute=0, second=0, microsecond=0) + timedelta(days=days_until_monday)
            wait_sec = (next_run - now).total_seconds()
            print(f"  [Prospects] Prochain scan : {next_run.strftime('%A %d/%m/%Y à %H:%M')} (dans {int(wait_sec//3600)}h)")
            time.sleep(wait_sec)
            print(f"  [Prospects] Scan hebdomadaire démarré — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            run_scan()  # Lit BCE_USER/BCE_PASSWORD depuis l'environnement

    t = threading.Thread(target=_loop, daemon=True)
    t.start()
    return t


def run_scan(credentials=None):
    """Lance un scan BCE complet et sauvegarde les résultats.
    Les credentials sont lus depuis les variables d'environnement Render :
    BCE_USER et BCE_PASSWORD — jamais en dur dans le code.
    """
    if not credentials:
        user = os.environ.get("BCE_USER", "")
        pwd  = os.environ.get("BCE_PASSWORD", "")
        if user and pwd:
            credentials = {"user": user, "password": pwd}
    try:
        result = download_and_parse_bce_update(credentials)
        if not result["ok"]:
            print(f"  [Prospects] Scan échoué : {result['error']}")
            return result
        summary = add_scan_result(result, enrich_emails=True)
        print(f"  [Prospects] Scan OK — {summary['new']} nouveaux, {summary['total']} total")
        return {"ok": True, **summary}
    except Exception as e:
        print(f"  [Prospects] Erreur : {e}")
        return {"ok": False, "error": str(e)[:100]}


# ── Export Excel (openpyxl si disponible, CSV sinon) ──────────
def export_prospects_excel(prospects, output_path):
    """Export les prospects en fichier Excel ou CSV."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prospects Dentaires"

        headers = [
            "Numéro BCE", "Dénomination", "Adresse", "CP", "Ville",
            "Date création", "Forme juridique", "NACE", "Téléphone",
            "Email", "Site Web", "Statut contact", "Date scan"
        ]

        # En-tête avec style Dental Addict
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="003087")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        # Données
        for row_idx, p in enumerate(prospects, 2):
            values = [
                f"BE{p.get('bce','')}",
                p.get("nom",""),
                p.get("adresse",""),
                p.get("cp",""),
                p.get("ville",""),
                p.get("date_creation",""),
                p.get("forme_juridique",""),
                ", ".join(p.get("nace_codes",[])),
                p.get("telephone",""),
                p.get("email",""),
                p.get("website",""),
                p.get("statut_contact","Nouveau"),
                p.get("scan_date",""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(vertical="center")
                # Alterner les couleurs de ligne
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="EBF0F9")

        # Largeurs de colonnes
        col_widths = [18,35,35,8,20,14,22,12,16,30,35,16,12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Filtre automatique
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

        # Figer la première ligne
        ws.freeze_panes = "A2"

        wb.save(output_path)
        return True, "xlsx"

    except ImportError:
        # Fallback CSV si openpyxl pas dispo
        with open(output_path.replace(".xlsx", ".csv"), "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(["Numéro BCE","Dénomination","Adresse","CP","Ville",
                             "Date création","Forme juridique","NACE","Téléphone",
                             "Email","Site Web","Statut contact","Date scan"])
            for p in prospects:
                writer.writerow([
                    f"BE{p.get('bce','')}", p.get("nom",""), p.get("adresse",""),
                    p.get("cp",""), p.get("ville",""), p.get("date_creation",""),
                    p.get("forme_juridique",""), ", ".join(p.get("nace_codes",[])),
                    p.get("telephone",""), p.get("email",""), p.get("website",""),
                    p.get("statut_contact","Nouveau"), p.get("scan_date","")
                ])
        return True, "csv"


if __name__ == "__main__":
    # Test standalone
    print("=== Test scan BCE NACE 8623 ===")
    result = download_and_parse_bce_update()
    print(json.dumps(result, indent=2, ensure_ascii=False)[:2000])
